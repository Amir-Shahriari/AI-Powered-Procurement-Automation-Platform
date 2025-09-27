# app/services/supplier.py
from __future__ import annotations

import json
import re
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Iterable

from fastapi import UploadFile

from ..config import settings
from app.services.textio import extract_text, chunks
from app.services.vectorstores import build_index, get_vs, _faiss_path, _chroma_name
from app.services.returnables_supplier import fill_returnables_from_supplier
from app.services.llm import rag_json

# Excel generation
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


# =============================================================================
# Paths & IO helpers
# =============================================================================

def _data_dir() -> Path:
    return settings.DATA_DIR

def _path_record_returnables(rec_id: str) -> Path:
    return _data_dir() / f"{rec_id}.returnables.json"

def _path_record_tepp(rec_id: str) -> Path:
    return _data_dir() / f"{rec_id}.tepp.json"

def _path_suppliers_manifest(rec_id: str) -> Path:
    return _data_dir() / f"{rec_id}.suppliers.json"

def _path_supplier_returnables(rec_id: str, sup_id: str) -> Path:
    return _data_dir() / f"{rec_id}_{sup_id}.returnables.json"

def _read_json(path: Path) -> Dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))

def _write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


# =============================================================================
# Supplier manifest management
# =============================================================================

def _load_suppliers(rec_id: str) -> List[Dict[str, Any]]:
    path = _path_suppliers_manifest(rec_id)
    if not path.exists():
        return []
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return []

def _save_suppliers(rec_id: str, items: List[Dict[str, Any]]) -> None:
    _path_suppliers_manifest(rec_id).write_text(json.dumps(items, indent=2), encoding="utf-8")

def list_suppliers(rec_id: str) -> List[Dict[str, Any]]:
    """
    Public: return the suppliers manifest (id, name, filename, uploaded_at, paths).
    If the manifest is empty/missing, try an on-disk rebuild so the UI doesn’t look empty.
    """
    items = _load_suppliers(rec_id)
    if not items:
        items = _rebuild_suppliers_manifest_from_disk(rec_id)
        if items:
            _save_suppliers(rec_id, items)
    return items


# =============================================================================
# Name extraction from Returnables
# =============================================================================

def _first_nonempty(*vals) -> Optional[str]:
    for v in vals:
        if isinstance(v, str) and v.strip():
            return v.strip()
    return None

def _get_nested(d: Dict[str, Any], path: List[str]) -> Any:
    cur = d
    for p in path:
        if isinstance(cur, dict):
            cur = cur.get(p)
        else:
            return None
    return cur

def _as_string(value: Any) -> Optional[str]:
    """
    Extract a human string from heterogeneous returnables field values.
    Common shapes:
      - "Acme Pty Ltd"
      - {"value": "Acme Pty Ltd", ...}
      - {"text": "Acme Pty Ltd"}
      - {"answer": "Acme Pty Ltd"}
      - [{"value": "Acme Pty Ltd"}, ...]
    """
    if isinstance(value, str):
        return value
    if isinstance(value, dict):
        # Prefer common payload keys, including nested under "value"
        prefer_keys = ["value", "text", "answer", "response", "selected", "name"]
        for k in prefer_keys:
            if k in value:
                s = _as_string(value.get(k))
                if s:
                    return s
        # Avoid label-like keys
        avoid_keys = {"label", "title", "question", "prompt", "heading", "field", "key"}
        for k, v in value.items():
            if k in avoid_keys:
                continue
            s = _as_string(v)
            if s:
                return s
    if isinstance(value, list):
        for el in value:
            s = _as_string(el)
            if s:
                return s
    return None

def _is_generic_label(s: str) -> bool:
    s_l = (s or "").strip().lower()
    if not s_l:
        return True
    generic = {
        "name","company name","organisation name","organization name","business name",
        "legal name","trading name","details legal name","details trading name",
        "enter company name","enter legal name"
    }
    # Check for example/placeholder text patterns
    # Only match these if they appear to be standalone or with minimal context
    if any(phrase in s_l for phrase in ["e.g.", "e.g.,", "example", "sample", "placeholder", "your company", "eg.", "eg,"]):
        return True
    
    # If it's a generic label or starts with common prefixes
    if s_l in generic or s_l.startswith("details ") or s_l.startswith("enter "):
        return True
    
    # Check for standalone business entity types (without a proper company name)
    standalone_patterns = ["pty ltd", "ltd", "limited", "inc", "llc", "corp", "company name", "business name"]
    if s_l in standalone_patterns:
        return True
    
    # Check for specific example patterns like "e.g., Pty Ltd"
    if re.match(r"^e\.?g\.?,?\s+(pty|ltd|limited|inc|corp|llc)", s_l, re.IGNORECASE):
        return True
    
    # If it's very short and looks like a placeholder
    if len(s_l) < 8 and any(standalone in s_l for standalone in ["company name", "business name", "legal name"]):
        return True
        
    return False

def _looks_like_company_name(s: str) -> bool:
    if not isinstance(s, str):
        return False
    s_clean = re.sub(r"\s+", " ", s).strip()
    if not s_clean:
        return False
    
    # Reject obvious placeholder/example text FIRST, before any other validation
    if _is_generic_label(s_clean):
        return False
    
    # Must be at least 5 characters long
    if len(s_clean) < 5:
        return False
    
    # Must contain at least one letter
    if not re.search(r'[A-Za-z]', s_clean):
        return False
    
    # Check for company suffixes - this is a strong indicator
    if re.search(r"\b(pty|pte|ltd|limited|inc|llc|llp|gmbh|s\.a\.|s\.r\.l\.|plc|corp|corporation)\b", s_clean, re.IGNORECASE):
        # Extract the part before the suffix
        before_suffix = re.split(r'\b(?:pty|pte|ltd|limited|inc|llc|llp|gmbh|s\.a\.|s\.r\.l\.|plc|corp|corporation)\b', s_clean, flags=re.IGNORECASE)[0].strip()
        # Additional check: make sure the part before suffix isn't generic either
        if _is_generic_label(before_suffix):
            return False
        # Must have at least 1 substantial word before the suffix (relaxed from 2)
        words_before = [w for w in re.split(r'[^A-Za-z]+', before_suffix) if w and len(w) >= 2]
        if len(words_before) >= 1 and len(before_suffix) >= 3:
            return True
    
    # For names without suffixes, require at least 2 substantial words (relaxed from 3)
    words = [w for w in re.split(r"[^A-Za-z]+", s_clean) if w and len(w) >= 2]
    return len(words) >= 2 and len(s_clean) >= 8

def _supplier_display_name(ret: Dict[str, Any], fallback: str) -> str:
    """
    Extract company name from populated returnables JSON.
    The returnables should already be populated by fill_returnables_from_supplier.
    """
    print(f"DEBUG: Extracting name from populated returnables, fallback: {fallback}")
    
    # First, try to get the company name from Schedule 1 - Corporate Information & Declaration
    schedule1 = _get_nested(ret, ["schedule_1_corporate_information_and_declaration"])
    if isinstance(schedule1, dict):
        print(f"DEBUG: Found Schedule 1 with keys: {list(schedule1.keys())}")
        
        # Check company_details first
        company_details = schedule1.get("company_details", {})
        if isinstance(company_details, dict):
            print(f"DEBUG: Company details keys: {list(company_details.keys())}")
            
            # Priority order for company name fields
            name_fields = ["legal_name", "trading_name", "company_name", "organisation_name", "organization_name", "business_name"]
            
            for field in name_fields:
                value = company_details.get(field)
                if value:
                    # Extract string value from potentially complex structure
                    name = _as_string(value)
                    print(f"DEBUG: {field} -> {value} -> extracted: {name}")
                    
                    if name and not _is_generic_label(name) and _looks_like_company_name(name):
                        print(f"DEBUG: Found valid company name in {field}: {name}")
                        return name.strip()
        
        # Check contact_details as fallback
        contact_details = schedule1.get("contact_details", {})
        if isinstance(contact_details, dict):
            print(f"DEBUG: Contact details keys: {list(contact_details.keys())}")
            
            contact_name_fields = ["company_name", "organisation_name", "organization_name", "name"]
            for field in contact_name_fields:
                value = contact_details.get(field)
                if value:
                    name = _as_string(value)
                    print(f"DEBUG: Contact {field} -> {value} -> extracted: {name}")
                    
                    if name and not _is_generic_label(name) and _looks_like_company_name(name):
                        print(f"DEBUG: Found valid company name in contact {field}: {name}")
                        return name.strip()
    
    # If Schedule 1 didn't work, try other common locations
    other_paths = [
        ["schedule_16_acknowledgement_of_addenda_and_declaration", "tenderer_declaration", "tenderer_name"],
        ["schedule_16_acknowledgement_of_addenda_and_declaration", "tenderer_declaration", "company_name"],
        ["company_name"],
        ["organization_name"],
        ["organisation_name"],
        ["legal_name"],
        ["trading_name"],
    ]
    
    for path in other_paths:
        value = _get_nested(ret, path)
        if value:
            name = _as_string(value)
            print(f"DEBUG: Path {path} -> {value} -> extracted: {name}")
            
            if name and not _is_generic_label(name) and _looks_like_company_name(name):
                print(f"DEBUG: Found valid company name at {path}: {name}")
                return name.strip()
    
    # Last resort: extract from all text content
    print("DEBUG: No structured name found, trying text extraction")
    text_content = _extract_text_from_returnables(ret)
    if text_content:
        # Look for company names with business suffixes using improved validation
        company_pattern = r'\b([A-Z][A-Za-z\s&.,\-]{4,50}?\s+(?:PTE?\.?|LTD\.?|LIMITED|INC\.?|INCORPORATED|CORP\.?|CORPORATION))\b'
        matches = re.findall(company_pattern, text_content, re.IGNORECASE)
        for match in matches:
            cleaned = match.strip()
            print(f"DEBUG: Found potential company name in text: {cleaned}")
            if (len(cleaned) > 3 and len(cleaned) < 100 and 
                not _is_generic_label(cleaned) and
                _looks_like_company_name(cleaned)):
                print(f"DEBUG: Validated company name from text: {cleaned}")
                return cleaned
            else:
                print(f"DEBUG: Rejected company name from text: {cleaned} (generic: {_is_generic_label(cleaned)}, looks_like: {_looks_like_company_name(cleaned)})")
    
    print(f"DEBUG: No company name found, using fallback: {fallback}")
    return fallback

def _extract_text_from_returnables(ret: Dict[str, Any]) -> str:
    """Extract all text content from returnables for pattern matching."""
    text_parts = []
    
    def extract_text_recursive(obj):
        if isinstance(obj, dict):
            for value in obj.values():
                extract_text_recursive(value)
        elif isinstance(obj, list):
            for item in obj:
                extract_text_recursive(item)
        elif isinstance(obj, str):
            text_parts.append(obj)
    
    extract_text_recursive(ret)
    return ' '.join(text_parts)

def _extract_company_name_from_raw_text(raw_path: Path) -> Optional[str]:
    """Extract company name directly from raw document text, focusing on Schedule 1."""
    try:
        text = extract_text(raw_path)
        if not text:
            return None
        
        print(f"DEBUG: Raw text length: {len(text)}")
        print(f"DEBUG: First 500 chars: {text[:500]}")
        
        def _schedule1_block(t: str) -> Optional[str]:
            # Find Schedule 1 section and slice until the next schedule header
            # Look for "Schedule 1" specifically (not 10, 11, etc.)
            m1 = re.search(r"(?is)\bSchedule\s*1(?:\s|$|[^\d])[\s\S]*?(?=\bSchedule\s*[2-9]|$)", t)
            if m1:
                block = m1.group(0)
                print(f"DEBUG: Found Schedule 1 block, length: {len(block)}")
                print(f"DEBUG: Schedule 1 content: {block[:300]}")
                return block
            
            # Fallback: try to find "Schedule 1" anywhere with word boundaries
            m2 = re.search(r"(?is)\bSchedule\s*1\b[\s\S]*?(?=\bSchedule\s*(?:[2-9]|\d{2})|$)", t)
            if m2:
                block = m2.group(0)
                print(f"DEBUG: Found Schedule 1 block (fallback), length: {len(block)}")
                print(f"DEBUG: Schedule 1 content: {block[:300]}")
                return block
            
            print("DEBUG: No Schedule 1 block found")
            return None

        # Get Schedule 1 block
        schedule1_text = _schedule1_block(text)
        if not schedule1_text:
            print("DEBUG: No Schedule 1 block found")
            return None

        # Look for company names in Schedule 1 - more specific patterns
        patterns = [
            # Direct company name patterns (like "SDE Consultants Pty Ltd") - require at least 2 words before suffix
            r'\b([A-Z][A-Za-z\s&.,\-]{4,50}?\s+(?:PTE?\.?|LTD\.?|LIMITED|INC\.?|INCORPORATED|CORP\.?|CORPORATION|LLC|LLP|CO\.?|COMPANY))\b',
            # After "Trading name" or "Legal name" labels - require at least 2 words
            r'(?:Trading\s+name|Legal\s+name|Company\s+name|Organisation\s+name|Organization\s+name)[:\s]*([A-Z][A-Za-z\s&.,\-]{4,50}?)(?:\s|$)',
            # After "Company:" or "Name:" labels - require at least 2 words
            r'(?:Company|Name)[:\s]*([A-Z][A-Za-z\s&.,\-]{4,50}?)(?:\s|$)',
            # Standalone company names (capitalized words with business suffix) - require at least 2 words before suffix
            r'\b([A-Z][A-Za-z\s&.,\-]{4,40}?\s+(?:PTE?\.?|LTD\.?|LIMITED))\b',
        ]
        
        # Also try to extract from the full document (not just Schedule 1) as a fallback
        # Look in the document header/early content where company names often appear
        if not schedule1_text:
            print("DEBUG: No Schedule 1 found, trying full document patterns")
            # Look in first 2000 characters for company names
            header_text = text[:2000]
            for i, pattern in enumerate(patterns):
                print(f"DEBUG: Trying header pattern {i+1}: {pattern}")
                matches = re.findall(pattern, header_text, re.IGNORECASE | re.MULTILINE)
                print(f"DEBUG: Found {len(matches)} header matches: {matches}")
                
                for match in matches:
                    cleaned = match.strip()
                    cleaned = re.sub(r'^[:\-\s]+|[:\-\s]+$', '', cleaned)
                    print(f"DEBUG: Testing header match: {cleaned}")
                    
                    if (len(cleaned) > 3 and len(cleaned) < 100 and 
                        not _is_generic_label(cleaned) and
                        _looks_like_company_name(cleaned)):
                        print(f"DEBUG: Returning company name from header: {cleaned}")
                        return cleaned
                    else:
                        print(f"DEBUG: Rejected header match: {cleaned} (generic: {_is_generic_label(cleaned)}, looks_like: {_looks_like_company_name(cleaned)})")
            return None

        # Search within Schedule 1 block only
        for i, pattern in enumerate(patterns):
            print(f"DEBUG: Trying pattern {i+1}: {pattern}")
            matches = re.findall(pattern, schedule1_text, re.IGNORECASE | re.MULTILINE)
            print(f"DEBUG: Found {len(matches)} matches: {matches}")
            
            for match in matches:
                cleaned = match.strip()
                # Clean up common artifacts
                cleaned = re.sub(r'^[:\-\s]+|[:\-\s]+$', '', cleaned)
                
                if (len(cleaned) > 3 and len(cleaned) < 100 and 
                    not _is_generic_label(cleaned) and
                    _looks_like_company_name(cleaned)):
                    print(f"DEBUG: Returning company name: {cleaned}")
                    return cleaned

        print("DEBUG: No valid company name found in Schedule 1")
        return None
    except Exception as e:
        print(f"DEBUG: Exception in raw text extraction: {e}")
        return None


# =============================================================================
# Public accessors for per-supplier Returnables
# =============================================================================

def get_supplier_returnables(rec_id: str, sup_id: str) -> Dict[str, Any]:
    path = _path_supplier_returnables(rec_id, sup_id)
    if not path.exists():
        raise FileNotFoundError(f"Supplier returnables not found: {sup_id}")
    return _read_json(path)

def save_supplier_returnables(rec_id: str, sup_id: str, payload: Dict[str, Any]) -> None:
    _write_json(_path_supplier_returnables(rec_id, sup_id), payload)
    # Optional: refresh supplier name in manifest in case user edited Schedule 1
    try:
        manifest = _load_suppliers(rec_id)
        for s in manifest:
            if s.get("id") == sup_id:
                s["name"] = _supplier_display_name(payload, s.get("name") or s.get("filename") or sup_id)
                break
        _save_suppliers(rec_id, manifest)
    except Exception:
        pass


# =============================================================================
# Supplier processing (fill returnables from uploaded files)
# =============================================================================

def process_suppliers(rec_id: str, files: List[UploadFile], base_returnables: Dict[str, Any], input_type: str = "responses") -> List[Dict[str, Any]]:
    """
    Ingest supplier materials and persist a per-supplier Returnables JSON.

    input_type:
      - "responses": supplier provides general response docs; we RAG-fill a blank RS.
      - "filled_returnables": supplier provides their already-filled Returnable Schedule
        (JSON preferred; PDF/DOCX supported via targeted Schedule 1 extraction).
    """
    if not files:
        items = _load_suppliers(rec_id)
        if not items:
            items = _rebuild_suppliers_manifest_from_disk(rec_id)
            if items:
                _save_suppliers(rec_id, items)
        return items

    input_type = (input_type or "responses").strip().lower()
    if input_type not in {"responses", "filled_returnables"}:
        input_type = "responses"

    manifest = _load_suppliers(rec_id)

    for up in files:
        if not up or not up.filename:
            continue
        # avoid dupes on Streamlit re-runs
        try:
            if any((s.get("filename") or "") == up.filename for s in manifest):
                continue
        except Exception:
            pass

        sup_id = str(uuid.uuid4())[:8]
        suffix = Path(up.filename).suffix or ".bin"
        raw_path = _data_dir() / f"{rec_id}_supplier_{sup_id}{suffix}"

        # Persist raw upload
        with open(raw_path, "wb") as f:
            f.write(up.file.read())

        filled: Dict[str, Any]
        supplier_name: Optional[str] = None
        vs_loc: Optional[str] = None

        # ---------- PATH A: Supplier uploaded a filled Returnable Schedule ----------
        if input_type == "filled_returnables":
            if suffix.lower() == ".json":
                # Direct JSON returnables from supplier (best path)
                try:
                    filled = json.loads(Path(raw_path).read_text(encoding="utf-8"))
                except Exception:
                    filled = json.loads(json.dumps(base_returnables))  # blank fallback
                # name: strictly from Schedule 1 first, then fallback
                supplier_name = _supplier_display_name(filled, Path(up.filename).stem)
            else:
                # PDF/DOCX: build a small VS and extract ONLY the fields we need
                text = ""
                try:
                    text = extract_text(raw_path)
                except Exception:
                    text = ""
                if not text.strip():
                    filled = json.loads(json.dumps(base_returnables))
                    supplier_name = Path(up.filename).stem
                else:
                    chs = chunks(text)
                    build_index(f"{rec_id}_supplier_{sup_id}", chs)
                    vs = get_vs(f"{rec_id}_supplier_{sup_id}")

                    # Start from blank RS but only request the minimal fields
                    filled = json.loads(json.dumps(base_returnables))

                    # Targeted Schedule 1 extraction (no broad scan)
                    try:
                        schedule1 = rag_json(
                            vs,
                            prompt=(
                                "Extract ONLY Schedule 1 Corporate Information & Declaration fields "
                                "(company_details.legal_name, company_details.trading_name, contact_details.*). "
                                "Do NOT return placeholders; return nulls if unknown. Return STRICT JSON."
                            ),
                            field_queries=[
                                "Schedule 1 Corporate Information company name legal_name trading_name",
                                "Corporate Information Declaration contact details company",
                            ],
                            max_ctx=6,
                            min_chars=1200,
                        ) or {}
                    except Exception:
                        schedule1 = {}

                    # Merge schedule1 (if any) into filled structure
                    if isinstance(schedule1, dict) and schedule1:
                        filled.setdefault("schedule_1_corporate_information_and_declaration", {})
                        filled["schedule_1_corporate_information_and_declaration"].update(schedule1)

                    # Prefer structured name from Schedule 1; no generic placeholders
                    name_from_rs = _supplier_display_name(filled, Path(up.filename).stem)
                    if name_from_rs and not _is_generic_label(name_from_rs):
                        supplier_name = name_from_rs

                    vs_loc = (
                        str(_faiss_path(f"{rec_id}_supplier_{sup_id}"))
                        if settings.VECTOR_BACKEND == "faiss_gpu"
                        else _chroma_name(f"{rec_id}_supplier_{sup_id}")
                    )

        # ---------- PATH B: Supplier uploaded general response docs ----------
        if input_type == "responses":
            # Extract text, index, fill full RS
            text = ""
            try:
                text = extract_text(raw_path)
            except Exception:
                text = ""

            if not text.strip():
                filled = json.loads(json.dumps(base_returnables))
                supplier_name = Path(up.filename).stem
            else:
                chs = chunks(text)
                build_index(f"{rec_id}_supplier_{sup_id}", chs)
                vs = get_vs(f"{rec_id}_supplier_{sup_id}")

                filled = fill_returnables_from_supplier(base_returnables, [vs])

                # 1) Try strict structured name from Schedule 1
                name_from_rs = _supplier_display_name(filled, Path(up.filename).stem)
                if name_from_rs and not _is_generic_label(name_from_rs):
                    supplier_name = name_from_rs

                # 2) If still missing/generic, a tiny targeted RAG for company name
                if not supplier_name or _is_generic_label(supplier_name):
                    try:
                        company_extraction = rag_json(
                            vs,
                            prompt=(
                                "Return STRICT JSON {'company_name': <actual company name>} "
                                "for the Tenderer. Do NOT return example/placeholder text."
                            ),
                            field_queries=[
                                "Schedule 1 company details legal name",
                                "Schedule 1 company details trading name",
                                "tenderer company name",
                            ],
                            max_ctx=6,
                            min_chars=1000,
                        ) or {}
                        n = company_extraction.get("company_name") if isinstance(company_extraction, dict) else None
                        if isinstance(n, str) and n.strip() and not _is_generic_label(n):
                            supplier_name = n.strip()
                    except Exception:
                        pass

                # 3) Last resort: regex within Schedule 1 block
                if not supplier_name or _is_generic_label(supplier_name):
                    from_raw = _extract_company_name_from_raw_text(raw_path)
                    if from_raw and not _is_generic_label(from_raw):
                        supplier_name = from_raw

                vs_loc = (
                    str(_faiss_path(f"{rec_id}_supplier_{sup_id}"))
                    if settings.VECTOR_BACKEND == "faiss_gpu"
                    else _chroma_name(f"{rec_id}_supplier_{sup_id}")
                )

        # ---- FINAL GUARANTEE: never persist None/generic name ----
        if not supplier_name or _is_generic_label(supplier_name):
            supplier_name = Path(up.filename).stem

        # Save supplier-specific returnables
        out_path = _path_supplier_returnables(rec_id, sup_id)
        _write_json(out_path, filled)

        manifest.append({
            "id": sup_id,
            "name": supplier_name,
            "filename": up.filename,
            "uploaded_at": datetime.utcnow().isoformat(),
            "returnables_path": str(out_path),
            "raw_path": str(raw_path),
            "vs_backend": settings.VECTOR_BACKEND,
            "vs_location": vs_loc,
        })

    _save_suppliers(rec_id, manifest)
    return manifest


# =============================================================================
# Auto-discovery / manifest rebuild (robust)
# =============================================================================

_DISCOVERY_PATTERNS = (
    # Common/expected:
    "{rec_id}_*.returnables.json",
    # Alternate some projects emit:
    "{rec_id}_supplier_*.returnables.json",
)

def _iter_candidate_returnables(rec_id: str) -> Iterable[Path]:
    """
    Yield all returnables JSON paths under DATA_DIR for this record id,
    searching recursively and trying multiple naming patterns.
    """
    root = _data_dir()
    for pat in _DISCOVERY_PATTERNS:
        for p in root.rglob(pat.format(rec_id=rec_id)):
            # Ensure it's a file (not directory) and ends with the expected suffix
            if p.is_file() and p.name.endswith(".returnables.json"):
                yield p

_SUP_ID_REGEXES = (
    # <rec_id>_<supid>.returnables.json
    lambda rec_id: re.compile(rf"^{re.escape(rec_id)}_([^/\\]+?)\.returnables\.json$", re.IGNORECASE),
    # <rec_id>_supplier_<supid>.returnables.json
    lambda rec_id: re.compile(rf"^{re.escape(rec_id)}_supplier_([^/\\]+?)\.returnables\.json$", re.IGNORECASE),
)

def _extract_sup_id_from_name(rec_id: str, filename: str) -> Optional[str]:
    for rx in _SUP_ID_REGEXES:
        m = rx(rec_id).match(filename)
        if m:
            return m.group(1)
    # Fallback: last token before .returnables.json
    stem = filename.replace(".returnables.json", "")
    parts = stem.split("_")
    if len(parts) >= 2:
        return parts[-1]
    return None

def _rebuild_suppliers_manifest_from_disk(rec_id: str) -> List[Dict[str, Any]]:
    """
    Rebuild suppliers manifest by scanning the data dir (recursively) for any
    returnables belonging to this record id, inferring supplier id and name.
    """
    rebuilt: List[Dict[str, Any]] = []
    for ret_path in _iter_candidate_returnables(rec_id):
        sup_id = _extract_sup_id_from_name(rec_id, ret_path.name)
        if not sup_id:
            continue
        try:
            ret = _read_json(ret_path)
        except Exception:
            ret = {}
        name = _supplier_display_name(ret, sup_id)

        # Try to guess raw and VS paths (best effort)
        guessed_raw = None
        # Consider both <rec_id>_<supid>.* and <rec_id>_supplier_<supid>.*
        for raw in _data_dir().rglob(f"{rec_id}_supplier_{sup_id}.*"):
            guessed_raw = str(raw)
            break
        if not guessed_raw:
            for raw in _data_dir().rglob(f"{rec_id}_{sup_id}.*"):
                if raw.suffix.lower() not in (".json",):
                    guessed_raw = str(raw)
                    break

        if settings.VECTOR_BACKEND == "faiss_gpu":
            vs_loc = str(_faiss_path(f"{rec_id}_supplier_{sup_id}"))
        else:
            # For Chroma, collection name only (best effort)
            vs_loc = _chroma_name(f"{rec_id}_supplier_{sup_id}")

        rebuilt.append({
            "id": sup_id,
            "name": name,
            "filename": None,
            "uploaded_at": None,
            "returnables_path": str(ret_path),
            "raw_path": guessed_raw,
            "vs_backend": settings.VECTOR_BACKEND,
            "vs_location": vs_loc,
        })

    # Stable ordering for deterministic workbooks
    rebuilt.sort(key=lambda x: (x.get("name") or "").lower() or (x.get("id") or ""))
    return rebuilt


# =============================================================================
# Evaluation workbook
# =============================================================================

def _load_weights_from_tepp(rec_id: str) -> List[Tuple[str, float, str]]:
    """
    Returns list of (criterion, weight_percent, rationale) from TEPP JSON.
    If not found, returns a basic default.
    """
    path = _path_record_tepp(rec_id)
    if not path.exists():
        return [
            ("Price (Full Cost)", 40.0, ""),
            ("Relevant Experience", 20.0, ""),
            ("Capability", 20.0, ""),
            ("Management & Financial Capacity", 10.0, ""),
            ("WHS", 10.0, ""),
        ]
    try:
        tepp = _read_json(path)
        rows = tepp.get("tender_evaluation", {}) \
                   .get("evaluation_methodology", {}) \
                   .get("required_criteria_table", []) or []
        out: List[Tuple[str, float, str]] = []
        for r in rows:
            crit = (r.get("criterion") or "").strip()
            wtxt = (r.get("weight") or "").strip()
            rationale = r.get("rationale") or ""
            m = re.search(r"(-?\d+(?:\.\d+)?)", wtxt)
            if crit and m:
                out.append((crit, float(m.group(1)), rationale))
        if out:
            return out
    except Exception:
        pass
    # fallback
    return [
        ("Price (Full Cost)", 40.0, ""),
        ("Relevant Experience", 20.0, ""),
        ("Capability", 20.0, ""),
        ("Management & Financial Capacity", 10.0, ""),
        ("WHS", 10.0, ""),
    ]

def _parse_money(s: Any) -> Optional[float]:
    """Parse '12,345.67', '$12,345', '12345' into float. Returns None if not parseable."""
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return float(s)
    txt = str(s).strip()
    if not txt:
        return None
    # Keep digits, dot, comma, minus
    cleaned = re.sub(r"[^\d,.\-]", "", txt)
    # If both comma and dot appear, assume comma=thousands, dot=decimal
    if "," in cleaned and "." in cleaned:
        cleaned = cleaned.replace(",", "")
    else:
        # If only comma appears, treat it as thousands
        cleaned = cleaned.replace(",", "")
    try:
        return float(cleaned)
    except Exception:
        return None

def _total_incl_gst_from_returnables(ret: Dict[str, Any]) -> Optional[float]:
    """
    Try to read Schedule 10 'pricing_totals.total_incl_gst'.
    If missing, attempt to sum 'pricing_schedule_ahu_refurbishments' incl_gst.
    """
    tot = ret.get("schedule_10_fee_structure", {}).get("pricing_totals", {}).get("total_incl_gst")
    v = _parse_money(tot)
    if v is not None:
        return v
    # Try sum of line items
    rows = ret.get("schedule_10_fee_structure", {}).get("pricing_schedule_ahu_refurbishments", []) or []
    acc = 0.0
    any_val = False
    for r in rows:
        val = _parse_money(r.get("price_incl_gst"))
        if val is not None:
            any_val = True
            acc += val
    return acc if any_val else None


def _style_header(ws, row=1):
    fill = PatternFill("solid", fgColor="1f2937")
    white = Font(color="FFFFFF", bold=True)
    for cell in ws[row]:
        cell.font = white
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")
    thin = Side(style="thin", color="374151")
    for cell in ws[row]:
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

def _autosize(ws):
    widths: Dict[int, float] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            val = str(cell.value)
            widths[cell.column] = max(widths.get(cell.column, 10.0), len(val) + 2.0)
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(w, 48)

def _safe_name(name: str) -> str:
    return re.sub(r"[\[\]*/\\?:]", "_", name)[:28] or "Supplier"

def _build_price_sheet(wb: Workbook, suppliers: List[Dict[str, Any]], price_weight: float):
    ws = wb.create_sheet("Price")
    headers = ["Supplier", "Pc (Total $)", "Ps", "Pn", "Price Weight (%)", "Pw (Weighted)"]
    ws.append(headers)
    _style_header(ws, 1)

    # Data rows
    for s in suppliers:
        ws.append([s["name"], s.get("pc") or None, None, None, price_weight, None])

    n = len(suppliers)
    # Average Pc range
    if n > 0:
        pc_range = f"B2:B{n+1}"
        for i in range(2, n + 2):
            # Ps = 200 – (Pc / Pav × 100) ; Pav = AVERAGE(Pc)
            ws[f"C{i}"] = f"=200 - (B{i} / AVERAGE({pc_range}) * 100)"
            # Pn = Ps / 200
            ws[f"D{i}"] = f"=C{i} / 200"
            # Pw = Pn × Price weight
            ws[f"F{i}"] = f"=D{i} * E{i}"

    _autosize(ws)
    return ws

def _build_weights_sheet(wb: Workbook, weights: List[Tuple[str, float, str]]):
    ws = wb.create_sheet("Weights")
    ws.append(["Criterion", "Weight (%)", "Rationale"])
    _style_header(ws, 1)
    for crit, w, rationale in weights:
        ws.append([crit, w, rationale or ""])
    _autosize(ws)
    return ws

def _build_scores_sheet(wb: Workbook,
                        suppliers: List[Dict[str, Any]],
                        nonprice_weights: List[Tuple[str, float]],
                        price_sheet_name: str):
    """
    Build "Scores" sheet with:
      - row 1: headers
      - row 2: weights
      - rows 3..: supplier ratings input (0..5) for non-price, Price Pw linked from 'Price'
      - Total and Rank columns
    """
    ws = wb.create_sheet("Scores")
    headers = ["Supplier"] + [c for c, _ in nonprice_weights] + ["Price (Weighted)", "Total (%)", "Rank"]
    ws.append(headers)
    _style_header(ws, 1)

    # Weights row
    weight_row = ["Weight (%)"] + [w for _, w in nonprice_weights] + [None, None, None]
    ws.append(weight_row)

    # Supplier rows
    for idx, s in enumerate(suppliers, start=3):
        row = [s["name"]]
        # Placeholder input cells for non-price ratings (0..5)
        row += [None for _ in nonprice_weights]
        # Link to Price!F{?}
        price_row = idx - 1  # Price sheet data starts at row 2
        row += [f"='{price_sheet_name}'!F{price_row}"]
        row += [None, None]  # Total, Rank (filled after append)
        ws.append(row)

    # Formulas for totals and rank
    n = len(suppliers)
    np_count = len(nonprice_weights)
    if n > 0:
        for r in range(3, n + 3):
            # Sum of (rating/5 * weight) across non-price cols
            contribs = []
            for j in range(np_count):
                rating_col = get_column_letter(2 + j)  # B .. (B + np_count - 1)
                weight_cell = f"{get_column_letter(2 + j)}2"  # weight row 2
                contribs.append(f"IFERROR(({rating_col}{r}/5)*{weight_cell},0)")
            nonprice_sum = "+".join(contribs) if contribs else "0"
            price_w_col = get_column_letter(2 + np_count)  # right after non-price
            total_col = get_column_letter(3 + np_count)    # Total column
            rank_col = get_column_letter(4 + np_count)     # Rank column

            ws[f"{total_col}{r}"] = f"=ROUND({nonprice_sum} + {price_w_col}{r}, 2)"
        # Rank (descending)
        total_col = get_column_letter(3 + np_count)
        total_range = f"${total_col}$3:${total_col}${n+2}"
        for r in range(3, n + 3):
            ws[f"{get_column_letter(4 + np_count)}{r}"] = f"=RANK({total_col}{r}, {total_range}, 0)"

    # Format bits
    for c in range(2, 2 + np_count):
        ws[f"{get_column_letter(c)}2"].number_format = "0.0"
    for r in range(3, n + 3):
        ws[f"{get_column_letter(2 + np_count)}{r}"].number_format = "0.0"  # Price weighted
        ws[f"{get_column_letter(3 + np_count)}{r}"].number_format = "0.00"  # Total
    _autosize(ws)
    return ws


def _create_evaluation_workbook(rec_id: str,
                                supplier_records: List[Dict[str, Any]],
                                weights: List[Tuple[str, float, str]]) -> Path:
    """
    Build an Excel workbook with three sheets: Weights, Price, Scores.
    Returns output file path.
    """
    # Partition weights
    price_weight = 0.0
    nonprice: List[Tuple[str, float]] = []
    for crit, w, _ in weights:
        if crit.lower().startswith("price"):
            price_weight = w
        else:
            nonprice.append((crit, w))

    # Suppliers enriched with Pc (price total)
    enriched: List[Dict[str, Any]] = []
    for s in supplier_records:
        ret = _read_json(Path(s["returnables_path"]))
        pc = _total_incl_gst_from_returnables(ret)
        enriched.append({"id": s["id"], "name": s["name"], "pc": pc})

    wb = Workbook()
    # Default sheet "Sheet" will be replaced by Weights
    wb.remove(wb.active)

    ws_w = _build_weights_sheet(wb, weights)
    ws_p = _build_price_sheet(wb, enriched, price_weight)
    ws_s = _build_scores_sheet(wb, enriched, nonprice, ws_p.title)

    # Make Scores the first sheet
    wb._sheets = [ws_s, ws_p, ws_w]

    out_path = _data_dir() / f"{rec_id}_evaluation.xlsx"
    wb.save(out_path)
    return out_path


def generate_evaluation_for_suppliers(rec_id: str, supplier_ids: Optional[List[str]] = None) -> Path:
    """
    Create a multi-supplier evaluation workbook.
    - If supplier_ids is None, include all suppliers in the manifest.
    - Uses TEPP weights when available.
    - If manifest is empty/missing, auto-discover suppliers from disk (recursively).
    """
    manifest = _load_suppliers(rec_id)
    if not manifest:
        manifest = _rebuild_suppliers_manifest_from_disk(rec_id)
        if manifest:
            _save_suppliers(rec_id, manifest)

    if not manifest:
        # Give a clear, actionable error with search details
        patterns = [p.format(rec_id=rec_id) for p in _DISCOVERY_PATTERNS]
        raise FileNotFoundError(
            "No suppliers have been processed for this record. "
            f"Searched recursively under DATA_DIR='{_data_dir()}' "
            f"for patterns: {patterns}. If you just uploaded files, "
            "ensure the /records/{rec_id}/supplier_responses endpoint writes "
            "per-supplier returnables like '{rec_id}_{sup_id}.returnables.json'."
        )

    selected = manifest
    if supplier_ids:
        idset = set(supplier_ids)
        selected = [s for s in manifest if s.get("id") in idset]
        if not selected:
            raise ValueError("None of the specified supplier IDs matched the manifest.")

    weights = _load_weights_from_tepp(rec_id)
    return _create_evaluation_workbook(rec_id, selected, weights)
